from __future__ import annotations

import asyncio
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from re import search
from urllib.parse import parse_qs, urlparse
from typing import Any

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from pyperclip import paste
from pydantic import BaseModel

try:
    from aiomysql import connect
except ImportError:  # pragma: no cover
    connect = None

from ..storage.mysql_config import DEFAULT_MYSQL_CONFIG
from ..tools import Browser


class TypeCreateRequest(BaseModel):
    name: str


class UrlCreateRequest(BaseModel):
    url: str = ""
    urls_text: str = ""
    fetch_type: str


class CollectOneRequest(BaseModel):
    task_id: int


class CollectAllRequest(BaseModel):
    fetch_type: str


class StopBatchRequest(BaseModel):
    batch_id: int


class DbConfigRequest(BaseModel):
    mysql_host: str
    mysql_port: int = 3306
    mysql_user: str = "root"
    mysql_password: str = ""
    mysql_database: str = "douk_downloader"


class DouyinTokenSwitchRequest(BaseModel):
    method: str
    browser: str = ""


class CommentCenterRepository:
    def __init__(self, parameter):
        self.parameter = parameter
        self._config = self._resolve_mysql_config(parameter)

    @staticmethod
    def _resolve_mysql_config(parameter) -> dict[str, Any]:
        settings_data = {}
        settings = getattr(parameter, "settings", None)
        if settings and hasattr(settings, "read"):
            try:
                loaded = settings.read()
                if isinstance(loaded, dict):
                    settings_data = loaded
            except Exception:
                settings_data = {}

        cfg: dict[str, Any] = {}
        for key, default in DEFAULT_MYSQL_CONFIG.items():
            if hasattr(parameter, key):
                cfg[key] = getattr(parameter, key)
            else:
                cfg[key] = settings_data.get(key, default)
        return cfg

    def get_mysql_config(self) -> dict[str, Any]:
        return {
            "mysql_host": str(self._config.get("mysql_host", "")),
            "mysql_port": int(self._config.get("mysql_port", 3306)),
            "mysql_user": str(self._config.get("mysql_user", "root")),
            "mysql_password": str(self._config.get("mysql_password", "")),
            "mysql_database": str(self._config.get("mysql_database", "douk_downloader")),
        }

    def update_mysql_config(self, cfg: dict[str, Any]):
        self._config.update(
            {
                "mysql_host": str(cfg.get("mysql_host", self._config.get("mysql_host", ""))),
                "mysql_port": int(cfg.get("mysql_port", self._config.get("mysql_port", 3306))),
                "mysql_user": str(cfg.get("mysql_user", self._config.get("mysql_user", "root"))),
                "mysql_password": str(cfg.get("mysql_password", self._config.get("mysql_password", ""))),
                "mysql_database": str(cfg.get("mysql_database", self._config.get("mysql_database", "douk_downloader"))),
            }
        )

    def persist_mysql_config(self, cfg: dict[str, Any]):
        settings = getattr(self.parameter, "settings", None)
        if not settings or not hasattr(settings, "read") or not hasattr(settings, "update"):
            return
        data = settings.read()
        if not isinstance(data, dict):
            data = {}
        data.update(
            {
                "mysql_host": cfg["mysql_host"],
                "mysql_port": int(cfg["mysql_port"]),
                "mysql_user": cfg["mysql_user"],
                "mysql_password": cfg["mysql_password"],
                "mysql_database": cfg["mysql_database"],
            }
        )
        settings.update(data)

    async def test_connection(self):
        async with self.connection() as _db:
            return

    async def _create_database_if_not_exists(self):
        if connect is None:
            raise HTTPException(status_code=500, detail="未安装 aiomysql")
        try:
            db = await connect(
                host=self._config["mysql_host"],
                port=int(self._config["mysql_port"]),
                user=self._config["mysql_user"],
                password=self._config["mysql_password"],
                charset="utf8mb4",
                autocommit=True,
            )
        except RuntimeError as e:
            if "cryptography" in str(e):
                raise HTTPException(
                    status_code=500,
                    detail=(
                        "MySQL 认证需要 cryptography 依赖，请执行："
                        " `uv pip install --python .venv/bin/python cryptography`"
                    ),
                ) from e
            raise
        try:
            async with db.cursor() as cursor:
                await cursor.execute("SET sql_notes = 0;")
                await cursor.execute(
                    f"CREATE DATABASE IF NOT EXISTS `{self._config['mysql_database']}` "
                    "CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;"
                )
                await cursor.execute("SET sql_notes = 1;")
        finally:
            db.close()

    @asynccontextmanager
    async def connection(self):
        if connect is None:
            raise HTTPException(status_code=500, detail="未安装 aiomysql")
        await self._create_database_if_not_exists()
        try:
            db = await connect(
                host=self._config["mysql_host"],
                port=int(self._config["mysql_port"]),
                user=self._config["mysql_user"],
                password=self._config["mysql_password"],
                db=self._config["mysql_database"],
                charset="utf8mb4",
                autocommit=True,
            )
        except RuntimeError as e:
            if "cryptography" in str(e):
                raise HTTPException(
                    status_code=500,
                    detail=(
                        "MySQL 认证需要 cryptography 依赖，请执行："
                        " `uv pip install --python .venv/bin/python cryptography`"
                    ),
                ) from e
            raise
        try:
            yield db
        finally:
            db.close()

    async def ensure_tables(self):
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute("SET sql_notes = 0;")
                try:
                    await cursor.execute(
                        """
                        CREATE TABLE IF NOT EXISTS `comment_types` (
                            `id` BIGINT NOT NULL AUTO_INCREMENT,
                            `name` VARCHAR(64) NOT NULL,
                            `created_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                            PRIMARY KEY (`id`),
                            UNIQUE KEY `uniq_name` (`name`)
                        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                        """
                    )

                    await cursor.execute(
                        """
                        CREATE TABLE IF NOT EXISTS `comment_url_tasks` (
                            `id` BIGINT NOT NULL AUTO_INCREMENT,
                            `url` VARCHAR(2048) NOT NULL,
                            `detail_id` VARCHAR(64) DEFAULT '',
                            `fetch_type` VARCHAR(64) NOT NULL,
                            `status` VARCHAR(32) NOT NULL DEFAULT '未处理',
                            `last_error` LONGTEXT,
                            `created_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                            `updated_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                            `last_fetch_at` TIMESTAMP NULL DEFAULT NULL,
                            PRIMARY KEY (`id`),
                            KEY `idx_fetch_type` (`fetch_type`),
                            KEY `idx_status` (`status`),
                            KEY `idx_detail_id` (`detail_id`)
                        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                        """
                    )

                    await cursor.execute(
                        """
                        CREATE TABLE IF NOT EXISTS `comment_records` (
                            `id` BIGINT NOT NULL AUTO_INCREMENT,
                            `task_id` BIGINT NOT NULL,
                            `url` VARCHAR(2048) NOT NULL,
                            `detail_id` VARCHAR(64) NOT NULL,
                            `fetch_type` VARCHAR(64) NOT NULL,
                            `dedupe_key` CHAR(64) NOT NULL,
                            `cid` VARCHAR(128) NULL,
                            `comment_time` DATETIME NULL,
                            `comment_time_text` VARCHAR(64) DEFAULT '',
                            `nickname` VARCHAR(255) DEFAULT '',
                            `uid` VARCHAR(128) DEFAULT '',
                            `sec_uid` VARCHAR(255) DEFAULT '',
                            `ip_label` VARCHAR(255) DEFAULT '',
                            `text` LONGTEXT,
                            `digg_count` BIGINT DEFAULT 0,
                            `reply_comment_total` BIGINT DEFAULT 0,
                            `raw_json` LONGTEXT,
                            `created_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                            PRIMARY KEY (`id`),
                            UNIQUE KEY `uniq_dedupe_key` (`dedupe_key`),
                            KEY `idx_task_id` (`task_id`),
                            KEY `idx_fetch_type` (`fetch_type`),
                            KEY `idx_comment_time` (`comment_time`),
                            KEY `idx_detail_id` (`detail_id`)
                        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                        """
                    )
                finally:
                    await cursor.execute("SET sql_notes = 1;")

                # 兼容已经创建过旧版本表结构的场景
                await self._ensure_column(cursor, "comment_records", "dedupe_key", "CHAR(64) NOT NULL DEFAULT ''")
                await cursor.execute(
                    """
                    UPDATE `comment_url_tasks`
                    SET `status` = CASE `status`
                        WHEN 'pending' THEN '未处理'
                        WHEN 'ready' THEN '未处理'
                        WHEN 'running' THEN '处理中'
                        WHEN 'done' THEN '处理完成'
                        WHEN 'failed' THEN '处理失败'
                        ELSE `status`
                    END
                    WHERE `status` IN ('pending', 'ready', 'running', 'done', 'failed');
                    """
                )
                await cursor.execute(
                    """
                    UPDATE `comment_records`
                    SET `dedupe_key` = SHA2(
                        CONCAT(
                            IFNULL(`detail_id`, ''), '|',
                            IFNULL(`cid`, ''), '|',
                            IFNULL(`comment_time_text`, ''), '|',
                            IFNULL(`nickname`, ''), '|',
                            IFNULL(`text`, ''), '|',
                            `id`
                        ),
                        256
                    )
                    WHERE `dedupe_key` IS NULL OR `dedupe_key` = '';
                    """
                )
                await self._ensure_unique_index(cursor, "comment_records", "uniq_dedupe_key", "dedupe_key")

    async def _ensure_column(self, cursor, table: str, column: str, ddl: str):
        await cursor.execute(
            """
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s AND COLUMN_NAME=%s
            """,
            (self._config["mysql_database"], table, column),
        )
        row = await cursor.fetchone()
        if row and int(row[0]) == 0:
            await cursor.execute(f"ALTER TABLE `{table}` ADD COLUMN `{column}` {ddl};")

    async def _ensure_unique_index(self, cursor, table: str, index_name: str, column: str):
        await cursor.execute(
            """
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.STATISTICS
            WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s AND INDEX_NAME=%s
            """,
            (self._config["mysql_database"], table, index_name),
        )
        row = await cursor.fetchone()
        if row and int(row[0]) == 0:
            await cursor.execute(
                f"CREATE UNIQUE INDEX `{index_name}` ON `{table}` (`{column}`);"
            )

    async def list_types(self) -> list[str]:
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute("SELECT name FROM `comment_types` ORDER BY id DESC")
                rows = await cursor.fetchall()
        return [r[0] for r in rows]

    async def add_type(self, name: str):
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(
                    "INSERT IGNORE INTO `comment_types` (`name`) VALUES (%s)",
                    (name,),
                )

    async def delete_type(self, name: str):
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute("DELETE FROM `comment_types` WHERE name=%s", (name,))

    async def type_exists(self, name: str) -> bool:
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute("SELECT COUNT(*) FROM `comment_types` WHERE name=%s", (name,))
                row = await cursor.fetchone()
        return bool(row and int(row[0]) > 0)

    async def insert_url_task(self, url: str, fetch_type: str) -> tuple[int, bool]:
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(
                    "SELECT id FROM `comment_url_tasks` WHERE url=%s LIMIT 1",
                    (url,),
                )
                row = await cursor.fetchone()
                if row:
                    return int(row[0]), False

                await cursor.execute(
                    "INSERT INTO `comment_url_tasks` (`url`, `fetch_type`) VALUES (%s, %s)",
                    (url, fetch_type),
                )
                return int(cursor.lastrowid), True

    async def list_url_tasks(
        self,
        fetch_type: str = "",
        status: str = "",
        page: int = 1,
        page_size: int = 20,
    ) -> dict[str, Any]:
        page = max(1, int(page))
        page_size = max(1, min(int(page_size), 200))
        offset = (page - 1) * page_size
        sql = "SELECT id, url, detail_id, fetch_type, status, last_error, created_at, updated_at, last_fetch_at FROM `comment_url_tasks`"
        count_sql = "SELECT COUNT(*) FROM `comment_url_tasks`"
        params: list[Any] = []
        where: list[str] = []
        if fetch_type:
            where.append("fetch_type=%s")
            params.append(fetch_type)
        if status:
            where.append("status=%s")
            params.append(status)
        if where:
            where_clause = " WHERE " + " AND ".join(where)
            sql += where_clause
            count_sql += where_clause
        sql += " ORDER BY id DESC LIMIT %s OFFSET %s"
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(count_sql, tuple(params))
                total_row = await cursor.fetchone()
                total = int(total_row[0]) if total_row else 0
                query_params = tuple([*params, page_size, offset])
                await cursor.execute(sql, query_params)
                rows = await cursor.fetchall()
        return {
            "items": [
                {
                    "id": r[0],
                    "url": r[1],
                    "detail_id": r[2],
                    "fetch_type": r[3],
                    "status": r[4],
                    "last_error": r[5],
                    "created_at": str(r[6]) if r[6] else None,
                    "updated_at": str(r[7]) if r[7] else None,
                    "last_fetch_at": str(r[8]) if r[8] else None,
                }
                for r in rows
            ],
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    async def get_task(self, task_id: int) -> dict[str, Any] | None:
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(
                    "SELECT id, url, detail_id, fetch_type, status FROM `comment_url_tasks` WHERE id=%s",
                    (task_id,),
                )
                row = await cursor.fetchone()
        if not row:
            return None
        return {
            "id": row[0],
            "url": row[1],
            "detail_id": row[2],
            "fetch_type": row[3],
            "status": row[4],
        }

    async def update_task_resolved(self, task_id: int, detail_id: str):
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(
                    "UPDATE `comment_url_tasks` SET detail_id=%s WHERE id=%s",
                    (detail_id, task_id),
                )

    async def update_task_status(self, task_id: int, status: str, error: str | None = None):
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(
                    "UPDATE `comment_url_tasks` SET status=%s, last_error=%s, last_fetch_at=NOW() WHERE id=%s",
                    (status, error, task_id),
                )

    async def list_url_task_ids_by_type(self, fetch_type: str) -> list[int]:
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(
                    "SELECT id FROM `comment_url_tasks` WHERE fetch_type=%s ORDER BY id DESC",
                    (fetch_type,),
                )
                rows = await cursor.fetchall()
        return [int(r[0]) for r in rows]

    async def list_url_task_ids_ready_for_batch(self, fetch_type: str) -> list[int]:
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(
                    """
                    SELECT id
                    FROM `comment_url_tasks`
                    WHERE fetch_type=%s
                      AND status NOT IN ('等待中', '处理中')
                    ORDER BY id DESC
                    """,
                    (fetch_type,),
                )
                rows = await cursor.fetchall()
        return [int(r[0]) for r in rows]

    async def batch_update_task_status(self, task_ids: list[int], status: str, error: str | None = None):
        ids = [int(i) for i in task_ids if int(i) > 0]
        if not ids:
            return
        placeholders = ", ".join(["%s"] * len(ids))
        sql = (
            f"UPDATE `comment_url_tasks` SET status=%s, last_error=%s, last_fetch_at=NOW() "
            f"WHERE id IN ({placeholders})"
        )
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(sql, tuple([status, error, *ids]))

    async def list_task_type_stats(self) -> list[dict[str, Any]]:
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(
                    """
                    SELECT
                        fetch_type,
                        COUNT(*) AS total_count,
                        SUM(CASE WHEN status IN ('处理完成', '处理失败', '已终止') THEN 1 ELSE 0 END) AS processed_count,
                        SUM(CASE WHEN status='处理中' THEN 1 ELSE 0 END) AS running_count,
                        SUM(CASE WHEN status='等待中' THEN 1 ELSE 0 END) AS waiting_count
                    FROM `comment_url_tasks`
                    GROUP BY fetch_type
                    ORDER BY fetch_type ASC
                    """
                )
                rows = await cursor.fetchall()
        return [
            {
                "fetch_type": str(r[0] or ""),
                "total_count": int(r[1] or 0),
                "processed_count": int(r[2] or 0),
                "running_count": int(r[3] or 0),
                "waiting_count": int(r[4] or 0),
            }
            for r in rows
            if str(r[0] or "")
        ]

    async def list_comment_type_counts(self) -> dict[str, int]:
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(
                    """
                    SELECT fetch_type, COUNT(*) AS comment_count
                    FROM `comment_records`
                    GROUP BY fetch_type
                    """
                )
                rows = await cursor.fetchall()
        return {str(r[0] or ""): int(r[1] or 0) for r in rows if str(r[0] or "")}

    async def upsert_comments(
        self,
        task_id: int,
        url: str,
        detail_id: str,
        fetch_type: str,
        items: list[dict[str, Any]],
    ):
        async with self.connection() as db:
            async with db.cursor() as cursor:
                for item in items:
                    user = item.get("user") if isinstance(item.get("user"), dict) else {}
                    uid = (
                        item.get("uid")
                        or user.get("uid")
                        or user.get("user_id")
                        or ""
                    )
                    nickname = (
                        item.get("nickname")
                        or user.get("nickname")
                        or user.get("nick_name")
                        or ""
                    )
                    sec_uid = (
                        item.get("sec_uid")
                        or user.get("sec_uid")
                        or user.get("secUid")
                        or ""
                    )
                    ip_label = (
                        item.get("ip_label")
                        or item.get("ipLabel")
                        or user.get("ip_label")
                        or ""
                    )
                    text = item.get("text") or ""
                    comment_time_text = item.get("create_time") or ""
                    comment_time = self._parse_comment_time(comment_time_text)
                    dedupe_key = self._build_dedupe_key(detail_id, item)
                    await cursor.execute(
                        """
                        INSERT INTO `comment_records` (
                            task_id, url, detail_id, fetch_type, dedupe_key, cid, comment_time, comment_time_text,
                            nickname, uid, sec_uid, ip_label, text, digg_count, reply_comment_total, raw_json
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        AS new
                        ON DUPLICATE KEY UPDATE
                            task_id=new.task_id,
                            url=new.url,
                            fetch_type=new.fetch_type,
                            cid=new.cid,
                            comment_time=new.comment_time,
                            comment_time_text=new.comment_time_text,
                            nickname=new.nickname,
                            uid=new.uid,
                            sec_uid=new.sec_uid,
                            ip_label=new.ip_label,
                            text=new.text,
                            digg_count=new.digg_count,
                            reply_comment_total=new.reply_comment_total,
                            raw_json=new.raw_json
                        """,
                        (
                            task_id,
                            url,
                            detail_id,
                            fetch_type,
                            dedupe_key,
                            item.get("cid") or None,
                            comment_time,
                            comment_time_text,
                            str(nickname),
                            str(uid),
                            str(sec_uid),
                            str(ip_label),
                            text,
                            int(item.get("digg_count") or 0),
                            int(item.get("reply_comment_total") or 0),
                            str(item),
                        ),
                    )

    @staticmethod
    def _build_dedupe_key(detail_id: str, item: dict[str, Any]) -> str:
        seed = "|".join(
            [
                detail_id,
                str(item.get("cid") or ""),
                str(item.get("create_time") or ""),
                str(item.get("nickname") or ""),
                str(item.get("text") or ""),
            ]
        )
        return sha256(seed.encode("utf-8")).hexdigest()

    @staticmethod
    def _parse_comment_time(value: str | int | float | None):
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            ts = float(value)
            # 兼容毫秒时间戳
            if ts > 10_000_000_000:
                ts = ts / 1000.0
            try:
                return datetime.fromtimestamp(ts)
            except (ValueError, OSError):
                return None
        if not isinstance(value, str):
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
        return None

    async def list_comments(
        self,
        detail_id: str = "",
        fetch_type: str = "",
        days: int | None = None,
        keyword: str = "",
        regions: list[str] | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> dict[str, Any]:
        page = max(1, int(page))
        page_size = max(1, min(int(page_size), 200))
        offset = (page - 1) * page_size
        where_sql, params = self._build_comment_filters(
            detail_id=detail_id,
            fetch_type=fetch_type,
            days=days,
            keyword=keyword,
            regions=regions or [],
        )
        sql = (
            "SELECT id, detail_id, uid, nickname, ip_label, comment_time, comment_time_text, text "
            f"FROM `comment_records` WHERE 1=1 {where_sql} "
            "ORDER BY comment_time DESC, id DESC LIMIT %s OFFSET %s"
        )
        count_sql = f"SELECT COUNT(*) FROM `comment_records` WHERE 1=1 {where_sql}"

        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(count_sql, tuple(params))
                total_row = await cursor.fetchone()
                total = int(total_row[0]) if total_row else 0
                query_params = tuple([*params, page_size, offset])
                await cursor.execute(sql, query_params)
                rows = await cursor.fetchall()

        return {
            "items": [
                {
                    "id": r[0],
                    "detail_id": r[1],
                    "uid": r[2] or "",
                    "nickname": r[3] or "",
                    "ip_label": r[4] or "",
                    "comment_time": str(r[5]) if r[5] else None,
                    "comment_time_text": r[6] or "",
                    "comment_time_display": self._format_comment_time_display(r[5], r[6]),
                    "text": r[7] or "",
                }
                for r in rows
            ],
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    async def list_comments_for_export(
        self,
        detail_id: str = "",
        fetch_type: str = "",
        days: int | None = None,
        keyword: str = "",
        regions: list[str] | None = None,
    ) -> list[dict[str, Any]]:
        where_sql, params = self._build_comment_filters(
            detail_id=detail_id,
            fetch_type=fetch_type,
            days=days,
            keyword=keyword,
            regions=regions or [],
        )
        sql = (
            "SELECT uid, nickname, ip_label, comment_time, comment_time_text, text "
            f"FROM `comment_records` WHERE 1=1 {where_sql} "
            "ORDER BY comment_time DESC, id DESC"
        )
        async with self.connection() as db:
            async with db.cursor() as cursor:
                await cursor.execute(sql, tuple(params))
                rows = await cursor.fetchall()
        return [
            {
                "uid": r[0] or "",
                "nickname": r[1] or "",
                "ip_label": r[2] or "",
                "comment_time_display": self._format_comment_time_display(r[3], r[4]),
                "text": r[5] or "",
            }
            for r in rows
        ]

    @staticmethod
    def _format_comment_time_display(comment_time: Any, comment_time_text: Any) -> str:
        if comment_time:
            if isinstance(comment_time, datetime):
                return comment_time.strftime("%Y-%m-%d %H:%M:%S")
            return str(comment_time)
        parsed = CommentCenterRepository._parse_comment_time(comment_time_text)
        if parsed:
            return parsed.strftime("%Y-%m-%d %H:%M:%S")
        return str(comment_time_text or "")

    @staticmethod
    def _build_comment_filters(
        detail_id: str,
        fetch_type: str,
        days: int | None,
        keyword: str,
        regions: list[str],
    ) -> tuple[str, list[Any]]:
        where: list[str] = []
        params: list[Any] = []
        if detail_id:
            where.append("detail_id=%s")
            params.append(detail_id)
        if fetch_type:
            where.append("fetch_type=%s")
            params.append(fetch_type)
        if days and days > 0:
            begin = datetime.now() - timedelta(days=days)
            where.append("comment_time IS NOT NULL AND comment_time >= %s")
            params.append(begin)
        if keyword:
            where.append("text LIKE %s")
            params.append(f"%{keyword}%")
        region_list = [r.strip() for r in regions if r and r.strip()]
        if region_list:
            region_sql = []
            for region in region_list:
                region_sql.append("ip_label LIKE %s")
                params.append(f"{region}%")
            where.append("(" + " OR ".join(region_sql) + ")")
        return (" AND " + " AND ".join(where)) if where else "", params


class CommentCenterController:
    def __init__(self, api_server):
        self.api_server = api_server
        self.repo = CommentCenterRepository(api_server.parameter)

    @staticmethod
    def _mask_cookie(cookie: dict[str, Any] | str) -> str:
        if isinstance(cookie, dict):
            if not cookie:
                return "未设置"
            has_login = "sessionid_ss" in cookie
            return f"已设置（{len(cookie)} 项，{'已登录' if has_login else '未登录'}）"
        if isinstance(cookie, str) and cookie.strip():
            return "已设置（字符串）"
        return "未设置"

    def get_douyin_token_status(self) -> dict[str, Any]:
        parameter = self.api_server.parameter
        settings_data = {}
        settings = getattr(parameter, "settings", None)
        if settings and hasattr(settings, "read"):
            loaded = settings.read()
            if isinstance(loaded, dict):
                settings_data = loaded
        cookie_val = settings_data.get("cookie", {})
        browser_default = (
            getattr(parameter, "browser_info", {}).get("browser_name", "Chrome")
            if hasattr(parameter, "browser_info")
            else "Chrome"
        )
        return {
            "cookie_status": self._mask_cookie(cookie_val),
            "browser_default": browser_default,
            "browsers": list(Browser.SUPPORT_BROWSER.keys()),
        }

    async def switch_douyin_token(self, method: str, browser_name: str = "") -> dict[str, Any]:
        parameter = self.api_server.parameter
        cookie_object = getattr(parameter, "cookie_object", None)
        if cookie_object is None:
            raise HTTPException(status_code=500, detail="Cookie 模块不可用")

        mode = method.strip().lower()
        cookie_data: dict[str, Any] = {}
        if mode == "clipboard":
            raw = paste()
            if not cookie_object.validate_cookie_minimal(raw):
                raise HTTPException(status_code=400, detail="当前剪贴板不是有效的 Cookie")
            cookie_data = cookie_object.extract(raw, write=False, key="cookie", platform="抖音")
        elif mode == "browser":
            select = (browser_name or getattr(parameter, "browser_info", {}).get("browser_name", "Chrome")).strip()
            browser = Browser(parameter, cookie_object)
            cookie_data = browser.get(select, Browser.PLATFORM[False].domain)
            if not cookie_data:
                raise HTTPException(status_code=400, detail=f"浏览器读取失败: {select}")
        else:
            raise HTTPException(status_code=400, detail="method 仅支持 clipboard 或 browser")

        cookie_object.save_cookie(cookie_data, "cookie")
        parameter.set_cookie(cookie_data, "")
        parameter.set_headers_cookie()
        try:
            await parameter.update_params()
        except Exception:
            # Cookie 已持久化，不阻断请求；下次轮询会继续刷新参数
            pass

        return {
            "ok": True,
            "message": "抖音 Token 已切换并保存，下次启动仍可使用",
            "cookie_status": self._mask_cookie(cookie_data),
        }

    async def resolve_detail_id(self, task: dict[str, Any]) -> str:
        if task.get("detail_id"):
            return task["detail_id"]
        ids = await self.api_server.links.run(task["url"])
        detail_id = ids[0] if ids else self._extract_detail_id_from_url(task["url"])
        if not detail_id:
            raise HTTPException(status_code=400, detail=f"提取作品ID失败: {task['url']}")
        await self.repo.update_task_resolved(task["id"], detail_id)
        return detail_id

    @staticmethod
    def _extract_detail_id_from_url(url: str) -> str:
        try:
            parsed = urlparse(url)
            query = parse_qs(parsed.query)
        except Exception:
            parsed = None
            query = {}

        for key in ("modal_id", "item_id", "aweme_id"):
            value = query.get(key, [""])[0]
            if value and value.isdigit():
                return value

        target = parsed.path if parsed else url
        if m := search(r"/video/(\d+)", target or ""):
            return m.group(1)
        if m := search(r"/note/(\d+)", target or ""):
            return m.group(1)
        return ""

    async def process_task(self, task_id: int) -> dict[str, Any]:
        task = await self.repo.get_task(task_id)
        if not task:
            raise HTTPException(status_code=404, detail="URL任务不存在")

        try:
            detail_id = await self.resolve_detail_id(task)
            await self.repo.update_task_status(task_id, "处理中", None)
            data = await self.api_server.comment_handle_single(
                detail_id,
                source=True,
            )
            if data is None:
                data = []
            await self.repo.upsert_comments(
                task_id,
                task["url"],
                detail_id,
                task["fetch_type"],
                data,
            )
            await self.repo.update_task_status(task_id, "处理完成", None)
            return {
                "task_id": task_id,
                "detail_id": detail_id,
                "fetch_type": task["fetch_type"],
                "count": len(data),
            }
        except HTTPException:
            await self.repo.update_task_status(task_id, "处理失败", "采集失败")
            raise
        except asyncio.CancelledError:
            await self.repo.update_task_status(task_id, "已终止", "任务被手动终止")
            raise
        except Exception as e:
            await self.repo.update_task_status(task_id, "处理失败", str(e))
            raise HTTPException(status_code=500, detail=f"采集失败: {e}") from e


class CommentTaskDispatcher:
    def __init__(self, controller: CommentCenterController):
        self.controller = controller
        self.queue: asyncio.Queue[tuple[int, int | None]] = asyncio.Queue()
        self.queued_ids: set[int] = set()
        self.worker_task: asyncio.Task | None = None
        self.current_item: tuple[int, int | None] | None = None
        self.lock = asyncio.Lock()
        self.batch_seq = 0
        self.active_batch: dict[str, Any] | None = None

    async def ensure_worker(self):
        if self.worker_task is None or self.worker_task.done():
            self.worker_task = asyncio.create_task(self._worker_loop())

    async def enqueue(self, task_id: int, batch_id: int | None = None) -> bool:
        await self.ensure_worker()
        async with self.lock:
            if task_id in self.queued_ids:
                return False
            self.queued_ids.add(task_id)
            await self.queue.put((task_id, batch_id))
            return True

    async def create_batch(self, fetch_type: str, total: int) -> tuple[dict[str, Any] | None, str]:
        async with self.lock:
            if self.active_batch and self.active_batch.get("status") in {"running", "stopping"}:
                return None, "已有任务进行中"
            self.batch_seq += 1
            self.active_batch = {
                "batch_id": self.batch_seq,
                "fetch_type": fetch_type,
                "status": "running",
                "total_count": int(total),
                "processed_count": 0,
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            return dict(self.active_batch), ""

    async def get_active_batch(self) -> dict[str, Any] | None:
        async with self.lock:
            if not self.active_batch:
                return None
            return dict(self.active_batch)

    async def stop_active_batch(self, batch_id: int) -> tuple[bool, str]:
        running_task_id: int | None = None
        pending_task_ids: list[int] = []
        async with self.lock:
            if not self.active_batch:
                return False, "当前没有批量任务在执行"
            if int(self.active_batch.get("batch_id", 0)) != int(batch_id):
                return False, "批量任务已变更，请刷新后重试"

            self.active_batch["status"] = "stopping"
            self.active_batch["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            drained: list[tuple[int, int | None]] = []
            while not self.queue.empty():
                try:
                    item = self.queue.get_nowait()
                except asyncio.QueueEmpty:
                    break
                self.queue.task_done()
                task_id, item_batch_id = item
                if item_batch_id == batch_id:
                    pending_task_ids.append(task_id)
                    self.queued_ids.discard(task_id)
                else:
                    drained.append(item)
            for item in drained:
                await self.queue.put(item)

            if self.current_item and self.current_item[1] == batch_id:
                running_task_id = self.current_item[0]

        if pending_task_ids:
            await self.controller.repo.batch_update_task_status(pending_task_ids, "已终止", "批量任务被手动终止")

        if self.worker_task and not self.worker_task.done():
            self.worker_task.cancel()
            try:
                await self.worker_task
            except asyncio.CancelledError:
                pass

        async with self.lock:
            if self.active_batch and int(self.active_batch.get("batch_id", 0)) == int(batch_id):
                processed = int(self.active_batch.get("processed_count") or 0)
                processed += len(pending_task_ids)
                self.active_batch["processed_count"] = min(processed, int(self.active_batch.get("total_count") or 0))
                self.active_batch["status"] = "stopped"
                self.active_batch["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.active_batch = None

        await self.ensure_worker()
        if running_task_id:
            return True, "已终止当前批量任务（含正在执行任务）"
        return True, "已终止当前批量任务"

    async def mark_batch_task_done(self, batch_id: int | None):
        if batch_id is None:
            return
        async with self.lock:
            if not self.active_batch or int(self.active_batch.get("batch_id", 0)) != int(batch_id):
                return
            self.active_batch["processed_count"] = int(self.active_batch.get("processed_count") or 0) + 1
            self.active_batch["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if int(self.active_batch["processed_count"]) >= int(self.active_batch.get("total_count") or 0):
                self.active_batch["status"] = "done"
                self.active_batch = None

    async def close(self):
        if self.worker_task and not self.worker_task.done():
            self.worker_task.cancel()
            try:
                await self.worker_task
            except asyncio.CancelledError:
                pass

    async def _worker_loop(self):
        while True:
            task_id, batch_id = await self.queue.get()
            self.current_item = (task_id, batch_id)
            try:
                await self.controller.process_task(task_id)
            except Exception:
                # 任务状态已在 process_task 内更新，这里吞掉异常避免 worker 退出
                pass
            finally:
                async with self.lock:
                    self.queued_ids.discard(task_id)
                    self.current_item = None
                self.queue.task_done()
                await self.mark_batch_task_done(batch_id)


def _split_urls(req: UrlCreateRequest) -> list[str]:
    urls: list[str] = []
    if req.url.strip():
        urls.append(req.url.strip())
    if req.urls_text.strip():
        for line in req.urls_text.splitlines():
            s = line.strip()
            if s:
                urls.append(s)
    # 保序去重
    seen = set()
    result = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            result.append(u)
    return result


def register_comment_center_routes(app: FastAPI, api_server):
    controller = CommentCenterController(api_server)
    dispatcher = CommentTaskDispatcher(controller)
    initialized = False

    async def ensure_storage_ready():
        nonlocal initialized
        if initialized:
            return
        await controller.repo.ensure_tables()
        initialized = True

    def apply_db_config(cfg: dict[str, Any]):
        nonlocal initialized
        controller.repo.update_mysql_config(cfg)
        controller.repo.persist_mysql_config(cfg)
        initialized = False

    @app.on_event("shutdown")
    async def _shutdown_comment_center_dispatcher():
        await dispatcher.close()

    @app.get("/comment-center", tags=["评论中心"])
    async def comment_center_home():
        page = Path(__file__).resolve().parents[2] / "static" / "comment_center.html"
        if not page.exists():
            return HTMLResponse("<h1>comment_center.html not found</h1>", status_code=500)
        return HTMLResponse(page.read_text(encoding="utf-8"))

    @app.get("/comment-center/urls", response_class=HTMLResponse, tags=["评论中心"])
    async def comment_center_urls_page():
        return RedirectResponse(url="/comment-center?tab=urls")

    @app.get("/comment-center/comments", response_class=HTMLResponse, tags=["评论中心"])
    async def comment_center_comments_page():
        return RedirectResponse(url="/comment-center?tab=comments")

    @app.get("/comment-center/import", response_class=HTMLResponse, tags=["评论中心"])
    async def comment_center_import_page():
        return RedirectResponse(url="/comment-center?tab=import")

    @app.get("/comment-center/settings", response_class=HTMLResponse, tags=["评论中心"])
    async def comment_center_settings_page():
        return RedirectResponse(url="/comment-center?tab=settings")

    @app.get("/comment-center/api/types", tags=["评论中心"])
    async def list_types():
        await ensure_storage_ready()
        return {"data": await controller.repo.list_types()}

    @app.post("/comment-center/api/types", tags=["评论中心"])
    async def create_type(item: TypeCreateRequest):
        await ensure_storage_ready()
        name = item.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="类型不能为空")
        await controller.repo.add_type(name)
        return {"ok": True}

    @app.get("/comment-center/api/db-config", tags=["评论中心"])
    async def get_db_config():
        return controller.repo.get_mysql_config()

    @app.get("/comment-center/api/douyin-token/status", tags=["评论中心"])
    async def get_douyin_token_status():
        return controller.get_douyin_token_status()

    @app.post("/comment-center/api/douyin-token/switch", tags=["评论中心"])
    async def switch_douyin_token(item: DouyinTokenSwitchRequest):
        return await controller.switch_douyin_token(
            method=item.method,
            browser_name=item.browser,
        )

    @app.post("/comment-center/api/db-config", tags=["评论中心"])
    async def set_db_config(item: DbConfigRequest):
        cfg = item.model_dump()
        apply_db_config(cfg)
        return {"ok": True, "message": "数据库配置已更新"}

    @app.post("/comment-center/api/db-config/test", tags=["评论中心"])
    async def test_db_config(item: DbConfigRequest):
        cfg = item.model_dump()
        apply_db_config(cfg)
        try:
            await ensure_storage_ready()
            await controller.repo.test_connection()
            return {"ok": True, "message": "数据库连接测试成功"}
        except Exception as e:
            return {"ok": False, "message": f"数据库连接测试失败: {e}"}

    @app.delete("/comment-center/api/types", tags=["评论中心"])
    async def remove_type(name: str = Query(...)):
        await ensure_storage_ready()
        n = name.strip()
        if not n:
            raise HTTPException(status_code=400, detail="类型不能为空")
        await controller.repo.delete_type(n)
        return {"ok": True}

    @app.get("/comment-center/api/url-list", tags=["评论中心"])
    async def list_urls(
        fetch_type: str = Query(default=""),
        status: str = Query(default=""),
        page: int = Query(default=1),
        page_size: int = Query(default=20),
    ):
        await ensure_storage_ready()
        return await controller.repo.list_url_tasks(
            fetch_type=fetch_type.strip(),
            status=status.strip(),
            page=page,
            page_size=page_size,
        )

    @app.post("/comment-center/api/url-list", tags=["评论中心"])
    async def create_url(item: UrlCreateRequest):
        await ensure_storage_ready()
        fetch_type = item.fetch_type.strip()
        if not fetch_type:
            raise HTTPException(status_code=400, detail="获取类型不能为空")
        if not await controller.repo.type_exists(fetch_type):
            raise HTTPException(status_code=400, detail="类型不存在，请先新增类型")

        urls = _split_urls(item)
        if not urls:
            raise HTTPException(status_code=400, detail="URL 不能为空")

        inserted = 0
        duplicated = 0
        ids = []
        for url in urls:
            task_id, created = await controller.repo.insert_url_task(url, fetch_type)
            ids.append(task_id)
            if created:
                inserted += 1
            else:
                duplicated += 1
        return {"inserted": inserted, "duplicated": duplicated, "ids": ids}

    @app.post("/comment-center/api/collect-one", tags=["评论中心"])
    async def collect_one(req: CollectOneRequest):
        await ensure_storage_ready()
        task = await controller.repo.get_task(req.task_id)
        if not task:
            raise HTTPException(status_code=404, detail="URL任务不存在")
        if task["status"] in {"等待中", "处理中"}:
            return {"ok": True, "message": "任务已在队列中或处理中"}
        await controller.repo.update_task_status(req.task_id, "等待中", None)
        queued = await dispatcher.enqueue(req.task_id)
        return {"ok": True, "queued": queued, "message": "任务后台处理中"}

    @app.post("/comment-center/api/collect-all", tags=["评论中心"])
    async def collect_all(req: CollectAllRequest):
        await ensure_storage_ready()
        fetch_type = req.fetch_type.strip()
        if not fetch_type:
            raise HTTPException(status_code=400, detail="获取类型不能为空")
        active_batch = await dispatcher.get_active_batch()
        if active_batch:
            return {
                "ok": False,
                "conflict": True,
                "active_batch": active_batch,
                "message": "已有批量任务进行中",
            }

        task_ids = await controller.repo.list_url_task_ids_ready_for_batch(fetch_type)
        if not task_ids:
            return {"ok": False, "queued": 0, "skipped": 0, "message": "当前类型没有可执行的任务"}

        batch, err = await dispatcher.create_batch(fetch_type, len(task_ids))
        if not batch:
            return {"ok": False, "conflict": True, "message": err or "已有批量任务进行中"}

        await controller.repo.batch_update_task_status(task_ids, "等待中", None)
        queued = 0
        for task_id in task_ids:
            if await dispatcher.enqueue(task_id, batch_id=int(batch["batch_id"])):
                queued += 1
        skipped = max(0, len(task_ids) - queued)
        return {
            "ok": True,
            "queued": queued,
            "skipped": skipped,
            "batch": batch,
            "message": "批量任务已启动",
        }

    @app.post("/comment-center/api/collect-all/stop", tags=["评论中心"])
    async def stop_collect_all(req: StopBatchRequest):
        await ensure_storage_ready()
        ok, message = await dispatcher.stop_active_batch(req.batch_id)
        return {"ok": ok, "message": message}

    @app.get("/comment-center/api/task-manager", tags=["评论中心"])
    async def task_manager():
        await ensure_storage_ready()
        type_stats = await controller.repo.list_task_type_stats()
        comment_counts = await controller.repo.list_comment_type_counts()
        active_batch = await dispatcher.get_active_batch()

        rows = []
        for item in type_stats:
            fetch_type = item["fetch_type"]
            running_count = int(item["running_count"])
            waiting_count = int(item["waiting_count"])
            status = "空闲"
            if active_batch and active_batch.get("fetch_type") == fetch_type:
                status = "进行中"
            elif running_count > 0 or waiting_count > 0:
                status = "进行中"
            rows.append(
                {
                    "fetch_type": fetch_type,
                    "total_count": int(item["total_count"]),
                    "processed_count": int(item["processed_count"]),
                    "comment_count": int(comment_counts.get(fetch_type, 0)),
                    "status": status,
                }
            )
        return {"items": rows, "active_batch": active_batch}

    @app.get("/comment-center/api/comments", tags=["评论中心"])
    async def list_comments(
        detail_id: str = Query(default=""),
        fetch_type: str = Query(default=""),
        time_range: str = Query(default=""),
        keyword: str = Query(default=""),
        regions: str = Query(default=""),
        page: int = Query(default=1),
        page_size: int = Query(default=20),
    ):
        await ensure_storage_ready()
        days = None
        if time_range in {"1", "3", "7"}:
            days = int(time_range)
        region_list = [r.strip() for r in regions.split(",") if r.strip()]
        return await controller.repo.list_comments(
            detail_id=detail_id.strip(),
            fetch_type=fetch_type.strip(),
            days=days,
            keyword=keyword.strip(),
            regions=region_list,
            page=page,
            page_size=page_size,
        )

    @app.get("/comment-center/api/comments/export", tags=["评论中心"])
    async def export_comments(
        detail_id: str = Query(default=""),
        fetch_type: str = Query(default=""),
        time_range: str = Query(default=""),
        keyword: str = Query(default=""),
        regions: str = Query(default=""),
    ):
        await ensure_storage_ready()
        days = None
        if time_range in {"1", "3", "7"}:
            days = int(time_range)
        region_list = [r.strip() for r in regions.split(",") if r.strip()]
        rows = await controller.repo.list_comments_for_export(
            detail_id=detail_id.strip(),
            fetch_type=fetch_type.strip(),
            days=days,
            keyword=keyword.strip(),
            regions=region_list,
        )

        wb = Workbook()
        wb.remove(wb.active)
        chunk_size = 5000
        headers = ["UID", "用户昵称", "地区", "评论时间", "评论内容"]
        if not rows:
            ws = wb.create_sheet("Data_1")
            ws.append(headers)
        else:
            for idx in range(0, len(rows), chunk_size):
                chunk = rows[idx : idx + chunk_size]
                sheet_no = idx // chunk_size + 1
                ws = wb.create_sheet(f"Data_{sheet_no}")
                ws.append(headers)
                for row in chunk:
                    ws.append(
                        [
                            row["uid"],
                            row["nickname"],
                            row["ip_label"],
                            row["comment_time_display"],
                            row["text"],
                        ]
                    )

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"comments_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
